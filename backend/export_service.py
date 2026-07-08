from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def generate_pdf_roster(roster_data, filename="roster.pdf"):
    """Generate a PDF roster from data"""
    
    doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e293b'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    elements.append(Paragraph("📋 Pharmacy Weekly Roster", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Prepare data for table
    if not roster_data:
        elements.append(Paragraph("No roster data available", styles['Normal']))
        doc.build(elements)
        return filename
    
    # Group by date
    dates = sorted(set(d['date'] for d in roster_data))
    days = sorted(set(d['day'] for d in roster_data), key=lambda x: ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'].index(x))
    
    # Create table data
    table_data = [['Employee'] + days]
    
    employees = sorted(set(d['user_name'] for d in roster_data))
    for emp in employees:
        row = [emp]
        for day in days:
            shift = next((d for d in roster_data if d['user_name'] == emp and d['day'] == day), None)
            if shift:
                row.append(f"{shift['branch_name']}\n{shift['start_time']}-{shift['end_time']}")
            else:
                row.append("Off")
        table_data.append(row)
    
    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return filename

def generate_excel_roster(roster_data, filename="roster.xlsx"):
    """Generate an Excel roster from data"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Roster"
    
    # Title
    ws['A1'] = "PHARMACY WEEKLY ROSTER"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Employee', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Total Shifts']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Group by date
    if roster_data:
        employees = sorted(set(d['user_name'] for d in roster_data))
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        
        row_num = 5
        for emp in employees:
            row = [emp]
            total_shifts = 0
            for day in days:
                shift = next((d for d in roster_data if d['user_name'] == emp and d['day'] == day), None)
                if shift:
                    row.append(f"{shift['branch_name']}\n{shift['start_time']}-{shift['end_time']}")
                    total_shifts += 1
                else:
                    row.append("Off")
            row.append(total_shifts)
            
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if value == "Off":
                    cell.fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
            
            row_num += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            for row in range(1, row_num + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[chr(64 + col)].width = max(15, max_length + 2)
    
    wb.save(filename)
    return filename

def generate_employee_pdf(employee_name, shifts, filename="employee_schedule.pdf"):
    """Generate a PDF for a single employee"""
    
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1e293b'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    elements.append(Paragraph(f"📋 Weekly Schedule for {employee_name}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    if not shifts:
        elements.append(Paragraph("No shifts scheduled this week", styles['Normal']))
        doc.build(elements)
        return filename
    
    # Create table data
    table_data = [['Day', 'Branch', 'Start Time', 'End Time']]
    days_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    
    for day in days_order:
        shift = next((s for s in shifts if s['day'] == day), None)
        if shift:
            # Use 'branch_name' if available, otherwise 'branch'
            branch = shift.get('branch_name', shift.get('branch', 'Unknown'))
            table_data.append([day, branch, shift['start_time'], shift['end_time']])
        else:
            table_data.append([day, '—', '—', '—'])
    
    table = Table(table_data, colWidths=[1.2*inch, 2*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return filename

# Test function
if __name__ == "__main__":
    # Test data - using consistent field names
    test_data = [
        {"date": "2026-07-13", "day": "Monday", "user_name": "Kevin Ambani", "branch_name": "Mumias Retail", "start_time": "07:00", "end_time": "16:00"},
        {"date": "2026-07-14", "day": "Tuesday", "user_name": "Kevin Ambani", "branch_name": "Sabatia Retail", "start_time": "09:00", "end_time": "21:00"},
        {"date": "2026-07-13", "day": "Monday", "user_name": "Billy Kipsang", "branch_name": "Mumias Wholesale", "start_time": "08:00", "end_time": "17:00"},
    ]
    
    # Generate PDF
    pdf_file = generate_pdf_roster(test_data, "test_roster.pdf")
    print(f"✅ PDF generated: {pdf_file}")
    
    # Generate Excel
    excel_file = generate_excel_roster(test_data, "test_roster.xlsx")
    print(f"✅ Excel generated: {excel_file}")
    
    # Generate employee PDF - use branch_name consistently
    employee_shifts = []
    for s in test_data:
        if s['user_name'] == "Kevin Ambani":
            employee_shifts.append({
                "day": s['day'],
                "branch_name": s['branch_name'],
                "start_time": s['start_time'],
                "end_time": s['end_time']
            })
    
    emp_pdf = generate_employee_pdf("Kevin Ambani", employee_shifts, "kevin_schedule.pdf")
    print(f"✅ Employee PDF generated: {emp_pdf}")